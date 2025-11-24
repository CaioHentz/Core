from decimal import Decimal, InvalidOperation
from collections import defaultdict

from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from django.contrib.auth.models import User as DjangoUser
from django.contrib.auth import authenticate, login

from .models import Purchase, Sale, Stock, Product, Supplier, Customer


def index(request):    
    # Compute key metrics
    sales_qs = Sale.objects.all()
    purchases_qs = Purchase.objects.all()

    total_sales_value = sum((s.total for s in sales_qs), Decimal("0"))
    total_purchases_value = sum((p.total for p in purchases_qs), Decimal("0"))
    total_profit = total_sales_value - total_purchases_value

    # Aggregate quantities sold per product
    sold_qty_map = defaultdict(Decimal)
    sold_value_map = defaultdict(Decimal)
    for s in sales_qs:
        try:
            qty = (s.quantity or Decimal("0"))
            sold_qty_map[s.product] += qty
            sold_value_map[s.product] += (qty * (s.price or Decimal("0")))
        except Exception:
            pass

    if sold_qty_map:
        most_sold_name, most_sold_qty = max(sold_qty_map.items(), key=lambda kv: kv[1])
    else:
        most_sold_name, most_sold_qty = "N/A", Decimal("0")

    # Chart data: Top 5 products by value sold
    top_sorted = sorted(sold_value_map.items(), key=lambda kv: kv[1], reverse=True)[:5]
    top_labels = [name for name, _ in top_sorted]
    top_data = [float(v) for _, v in top_sorted]

    # Chart data: Totals (sales vs purchases)
    totals_labels = ["Sales", "Purchases"]
    totals_data = [float(total_sales_value), float(total_purchases_value)]

    # Formatting helpers: up to 3 decimals (trim trailing zeros) with thousands '.' and decimal ','
    def _fmt_number(dec):
        dec = dec or Decimal("0")
        q = dec.quantize(Decimal("0.001"))
        s = format(q, "f")
        neg = s.startswith("-")
        if "." in s:
            int_part, frac_part = s.split(".")
            frac_part = frac_part.rstrip("0")
        else:
            int_part, frac_part = s, ""
        # remove sign from integer part for grouping
        if neg:
            int_part = int_part.lstrip("-")
        # group thousands with '.'
        int_rev = int_part[::-1]
        grouped_rev = ".".join(int_rev[i:i+3] for i in range(0, len(int_rev), 3))
        grouped = grouped_rev[::-1]
        res = f"{grouped},{frac_part}" if frac_part else grouped
        # avoid '-0' representation
        return f"-{res}" if neg and res != "0" else res

    def _currency(dec):
        return f"${_fmt_number(dec)}"

    context = {
        "metrics": {
            "total_sales": _currency(total_sales_value),
            "total_purchases": _currency(total_purchases_value),
            "total_profit": _currency(total_profit),
            "product_most_sold": most_sold_name,
            "product_most_sold_qty": _fmt_number(most_sold_qty),
        },
        "charts": {
            "topProducts": {"labels": top_labels, "data": top_data},
            "totals": {"labels": totals_labels, "data": totals_data},
        },
    }
    return render(request, "galeria/index.html", context)


def purchase(request):
    if request.method == "POST":
        supplier = (request.POST.get("supplier") or "").strip()
        product = (request.POST.get("product") or "").strip()
        quantity_str = (request.POST.get("quantity") or "").strip()
        price_str = (request.POST.get("price") or "").strip()

        if not supplier or not product or not quantity_str or not price_str:
            messages.error(request, "All fields are required.")
        else:
            try:
                qty = Decimal(quantity_str)
                price = Decimal(price_str)
                if qty <= 0 or price < 0:
                    messages.error(request, "Quantity must be > 0 and price must be >= 0.")
                else:
                    # Validate supplier and product existence before recording purchase
                    if not Supplier.objects.filter(name__iexact=supplier).exists():
                        messages.error(request, f"Supplier '{supplier}' does not exist. Please create it first in Suppliers.")
                    elif not Product.objects.filter(name__iexact=product).exists():
                        messages.error(request, f"Product '{product}' does not exist. Please create it first in Products.")
                    else:
                        Purchase.objects.create(
                            supplier=supplier,
                            product=product,
                            quantity=qty,
                            price=price,
                        )
                        # Update stock 
                        Stock.add_stock(product, qty)
                        messages.success(request, f"Added purchase: {qty} x '{product}' from '{supplier}' at {price}.")
                        return redirect("purchase")
            except InvalidOperation:
                messages.error(request, "Invalid number format for quantity or price.")

    purchases = Purchase.objects.order_by("-created_at")
    products_qs = Product.objects.order_by("name")
    suppliers_qs = Supplier.objects.order_by("name")
    uom_map = {p.name.lower(): p.unit_of_measure for p in products_qs}
    for obj in purchases:
        try:
            obj_uom = uom_map.get(obj.product.lower(), "")
        except AttributeError:
            obj_uom = ""
        setattr(obj, "uom", obj_uom)
    context = {
        "purchases": purchases,
        "products": products_qs,
        "suppliers": suppliers_qs,
        "uom_map": uom_map,
    }
    return render(request, "galeria/purchase.html", context)


def sales(request):
    if request.method == "POST":
        customer = (request.POST.get("customer") or "").strip()
        product = (request.POST.get("product") or "").strip()
        quantity_str = (request.POST.get("quantity") or "").strip()
        price_str = (request.POST.get("price") or "").strip()

        if not customer or not product or not quantity_str or not price_str:
            messages.error(request, "All fields are required.")
        else:
            try:
                qty = Decimal(quantity_str)
                price = Decimal(price_str)
                if qty <= 0 or price < 0:
                    messages.error(request, "Quantity must be > 0 and price must be >= 0.")
                else:
                    # Validate customer and product existence before sale
                    if not Customer.objects.filter(name__iexact=customer).exists():
                        messages.error(request, f"Customer '{customer}' does not exist. Please create it first in Customers.")
                    elif not Product.objects.filter(name__iexact=product).exists():
                        messages.error(request, f"Product '{product}' does not exist. Please create it first in Products.")
                    else:
                        # Attempt to remove stock before creating sale 
                        if Stock.remove_stock(product, qty):
                            Sale.objects.create(
                                customer=customer,
                                product=product,
                                quantity=qty,
                                price=price,
                            )
                            messages.success(request, f"Added sale: {qty} x '{product}' to '{customer}' at {price}.")
                            return redirect("sales")
                        else:
                            messages.error(request, f"Insufficient stock for '{product}'.")
            except InvalidOperation:
                messages.error(request, "Invalid number format for quantity or price.")

    sales_qs = Sale.objects.order_by("-created_at")
    products_qs = Product.objects.order_by("name")
    customers_qs = Customer.objects.order_by("name")
    uom_map = {p.name.lower(): p.unit_of_measure for p in products_qs}
    for obj in sales_qs:
        try:
            obj_uom = uom_map.get(obj.product.lower(), "")
        except AttributeError:
            obj_uom = ""
        setattr(obj, "uom", obj_uom)
    context = {
        "sales": sales_qs,
        "products": products_qs,
        "customers": customers_qs,
        "uom_map": uom_map,
    }
    return render(request, "galeria/sales.html", context)



def inventory(request):
    query = request.GET.get("q", "")

    stocks = Stock.objects.all().order_by("product")

    if query:
        stocks = stocks.filter(product__icontains=query)

    products_qs = Product.objects.all()
    uom_map = {p.name.lower(): p.unit_of_measure for p in products_qs}

    for obj in stocks:
        try:
            obj_uom = uom_map.get(obj.product.lower(), "")
        except AttributeError:
            obj_uom = ""
        setattr(obj, "uom", obj_uom)

    context = {
        "stocks": stocks,
        "query": query,
    }

    return render(request, "galeria/inventory.html", context)





def products(request):
    products_qs = Product.objects.order_by("name")
    context = {"products": products_qs}
    return render(request, "galeria/products.html", context)


def product_create(request):
    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        description = (request.POST.get("description") or "").strip()
        uom = (request.POST.get("unit_of_measure") or "").strip()

        if not name or not uom:
            messages.error(request, "Name and unit of measure are required.")
        elif Product.objects.filter(name__iexact=name).exists():
            messages.error(request, f"Product '{name}' already exists.")
        else:
            Product.objects.create(name=name, description=description, unit_of_measure=uom)
            messages.success(request, f"Product '{name}' created.")
            return redirect("products")

    return render(request, "galeria/product_form.html")


def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        description = (request.POST.get("description") or "").strip()
        uom = (request.POST.get("unit_of_measure") or "").strip()

        if not name or not uom:
            messages.error(request, "Name and unit of measure are required.")
        elif Product.objects.filter(name__iexact=name).exclude(pk=pk).exists():
            messages.error(request, f"Product '{name}' already exists.")
        else:
            product.name = name
            product.description = description
            product.unit_of_measure = uom
            product.save()
            messages.success(request, f"Product '{name}' updated.")
            return redirect("products")

    return render(request, "galeria/product_form.html", {"product": product})


def customers(request):
    customers_qs = Customer.objects.order_by("name")
    context = {"customers": customers_qs}
    return render(request, "galeria/customers.html", context)


def customer_create(request):
    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        description = (request.POST.get("description") or "").strip()

        if not name:
            messages.error(request, "Name is required.")
        elif Customer.objects.filter(name__iexact=name).exists():
            messages.error(request, f"Customer '{name}' already exists.")
        else:
            Customer.objects.create(name=name, description=description)
            messages.success(request, f"Customer '{name}' created.")
            return redirect("customers")

    return render(request, "galeria/customer_form.html")


def customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk)

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        description = (request.POST.get("description") or "").strip()

        if not name:
            messages.error(request, "Name is required.")
        elif Customer.objects.filter(name__iexact=name).exclude(pk=pk).exists():
            messages.error(request, f"Customer '{name}' already exists.")
        else:
            customer.name = name
            customer.description = description
            customer.save()
            messages.success(request, f"Customer '{name}' updated.")
            return redirect("customers")

    return render(request, "galeria/customer_form.html", {"customer": customer})


def suppliers(request):
    suppliers_qs = Supplier.objects.order_by("name")
    context = {"suppliers": suppliers_qs}
    return render(request, "galeria/suppliers.html", context)


def supplier_create(request):
    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        description = (request.POST.get("description") or "").strip()

        if not name:
            messages.error(request, "Name is required.")
        elif Supplier.objects.filter(name__iexact=name).exists():
            messages.error(request, f"Supplier '{name}' already exists.")
        else:
            Supplier.objects.create(name=name, description=description)
            messages.success(request, f"Supplier '{name}' created.")
            return redirect("suppliers")

    return render(request, "galeria/supplier_form.html")


def supplier_edit(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        description = (request.POST.get("description") or "").strip()

        if not name:
            messages.error(request, "Name is required.")
        elif Supplier.objects.filter(name__iexact=name).exclude(pk=pk).exists():
            messages.error(request, f"Supplier '{name}' already exists.")
        else:
            supplier.name = name
            supplier.description = description
            supplier.save()
            messages.success(request, f"Supplier '{name}' updated.")
            return redirect("suppliers")

    return render(request, "galeria/supplier_form.html", {"supplier": supplier})


# Export to Excel endpoints

def sales_export(request):
    sales_qs = Sale.objects.order_by("created_at")
    columns = ["Date", "Customer", "Product", "Quantity", "Price (Unit)", "Total"]
    rows = []
    for s in sales_qs:
        rows.append([
            s.created_at.strftime("%Y-%m-%d %H:%M"),
            s.customer,
            s.product,
            float(s.quantity or Decimal("0")),
            float(s.price or Decimal("0")),
            float((s.quantity or Decimal("0")) * (s.price or Decimal("0"))),
        ])
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(columns)
    for r in rows:
        ws.append(r)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="sales.xlsx"'
    return response


def purchase_export(request):
    purchases_qs = Purchase.objects.order_by("created_at")
    columns = ["Date", "Supplier", "Product", "Quantity", "Price (Unit)", "Total"]
    rows = []
    for p in purchases_qs:
        rows.append([
            p.created_at.strftime("%Y-%m-%d %H:%M"),
            p.supplier,
            p.product,
            float(p.quantity or Decimal("0")),
            float(p.price or Decimal("0")),
            float((p.quantity or Decimal("0")) * (p.price or Decimal("0"))),
        ])
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchases"
    ws.append(columns)
    for r in rows:
        ws.append(r)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="purchases.xlsx"'
    return response


def inventory_export(request):
    stocks = Stock.objects.order_by("product")
    products_qs = Product.objects.all()
    uom_map = {p.name.lower(): p.unit_of_measure for p in products_qs}
    columns = ["Product", "Quantity", "UoM"]
    rows = []
    for s in stocks:
        try:
            uom = uom_map.get(s.product.lower(), "")
        except AttributeError:
            uom = ""
        rows.append([
            s.product,
            float(s.quantity or Decimal("0")),
            uom,
        ])
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(columns)
    for r in rows:
        ws.append(r)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="inventory.xlsx"'
    return response

def register_account_view(request):
    if request.method == "POST":
        username = request.POST.get("username") or ""
        password = request.POST.get("password") or ""

        if not username or not password:
            messages.error(request, "You must fill all fields.")
            return redirect("register_account_view")

        if DjangoUser.objects.filter(username__iexact=username).exists():
            messages.error(request, "Username already exists.")
            return redirect("register_account_view")

        user = DjangoUser(username=username)
        user.set_password(password)
        user.save()

        messages.success(request, "Account created successfully! Please login.")
        return redirect("authenticate_account_view")

    return render(request, "galeria/register.html")

def authenticate_account_view(request):
    if request.method == "POST":
        username = request.POST.get("username") or ""
        password = request.POST.get("password") or ""

        if not username or not password:
            messages.error(request, "You must fill all fields.")
            return redirect("authenticate_account_view")

        user = authenticate(request, username=username, password=password)
        if user is None:
            messages.error(request, "Invalid username or password.")
            return redirect("authenticate_account_view")

        login(request, user)
        messages.success(request, "Login successful!")
        return redirect("index")

    return render(request, "galeria/login.html")
